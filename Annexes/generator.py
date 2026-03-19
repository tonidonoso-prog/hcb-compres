import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Protection, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.drawing.image import Image
import os

def get_column_mapping(wb_in, annex_type, lang='es'):
    import re
    def normalize(text):
        return set(re.findall(r'[A-Z0-9]+', str(text).upper()))

    try:
        import os
        fname = 'CABECERAS-cat.xlsx' if lang == 'cat' else 'CABECERAS.xlsx'
        cab_file = os.path.join(os.path.dirname(__file__), fname)
        if not os.path.exists(cab_file):
            cab_file = fname
            if not os.path.exists(cab_file):
                raise ValueError(f"No se ha encontrado el fichero de cabeceras: {fname}")

        df_map = pd.read_excel(cab_file, sheet_name=annex_type, header=None)
        annex_headers = [str(h).strip() for h in df_map.iloc[0].fillna("").tolist()]
        hi_headers_to_find = [str(h).strip().upper() for h in df_map.iloc[1].fillna("").tolist()]

        ws_in = wb_in['Full Inici']
        mapping = {}
        fuzzy_matches = []
        not_found = []

        for r_idx in [4, 5, 6, 7]:
            row_vals = [str(ws_in.cell(row=r_idx, column=c).value).strip().upper() for c in range(1, 150)]
            for ah, htf in zip(annex_headers, hi_headers_to_find):
                if ah in mapping: continue
                if htf == "NAN" or htf == "": continue

                if htf in row_vals:
                    mapping[ah] = row_vals.index(htf) + 1
                else:
                    best_idx = -1
                    best_score = 0
                    best_match_text = ""
                    htf_words = normalize(htf)
                    if not htf_words: continue

                    for idx, rv in enumerate(row_vals):
                        if rv == "" or rv == "NONE": continue
                        rv_words = normalize(rv)
                        if not rv_words: continue

                        if len(rv) >= 4 and len(htf) >= 4 and (htf in rv or rv in htf):
                            best_idx = idx + 1
                            best_score = 2.0
                            best_match_text = rv
                            break

                        intersection = htf_words.intersection(rv_words)
                        score = len(intersection) / max(len(htf_words), len(rv_words))
                        if score > best_score and score >= 0.5:
                            best_score = score
                            best_idx = idx + 1
                            best_match_text = rv

                    if best_score > 0:
                        mapping[ah] = best_idx
                        if best_score < 2.0:
                            fuzzy_matches.append(f"'{ah}' → fuzzy match '{best_match_text}' (score {best_score:.0%})")

        # Detectar columnas no encontradas
        for ah, htf in zip(annex_headers, hi_headers_to_find):
            if htf == "NAN" or htf == "": continue
            if ah not in mapping:
                not_found.append(ah)

        warnings = []
        if not_found:
            warnings.append(f"Columnas no encontradas en el HI: {', '.join(not_found)}")
        if fuzzy_matches:
            warnings.append("Columnas mapeadas por similitud (revisar): " + "; ".join(fuzzy_matches))

        return mapping, warnings
    except Exception as e:
        if isinstance(e, ValueError):
            raise
        raise ValueError(f"Error leyendo cabeceras ({annex_type}, {lang}): {e}")

def apply_style(ws, rng, text=None, fill=None, font=None, align=None, border=None):
    if not align:
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if ":" in rng: ws.merge_cells(rng)
    start_cell = ws[rng.split(":")[0]]
    if text is not None: start_cell.value = text
    start_coords, end_coords = rng.split(":") if ":" in rng else (rng, rng)
    for row in ws[start_coords:end_coords]:
        for cell in row:
            cell.alignment = align
            if border: cell.border = border
            if fill: cell.fill = fill
            if font: cell.font = font

def generate_am(input_bytes, logo_path='logo.png', lang='es'):
    wb_in = load_workbook(io.BytesIO(input_bytes), data_only=True)

    # --- DADES DE CABECERA ---
    nom_pestanya_cab = next((s for s in wb_in.sheetnames if s.upper() == "CABECERA"), None)
    if not nom_pestanya_cab:
        raise ValueError("No s'ha trobat la pestanya 'cabecera' al fitxer original.")

    ws_cab = wb_in[nom_pestanya_cab]
    titol_expedient = ws_cab['B9'].value
    num_expedient = ws_cab['B5'].value

    # --- TEXTOS PER IDIOMA ---
    if lang == 'cat':
        t = {
            'title': "ANNEX DE COMPLIMENTACIÓ OBLIGATÒRIA 2 PPT ALBARÀ DE MOSTRES",
            'instr': ("Haurà de presentar les mostres relacionades en aquest albarà segons s'indiqui als plecs.\n"
                      "Ompli només les cel·les en blanc o desprotegides."),
            'tit_exp': "TÍTOL DE L'EXPEDIENT:", 'num_exp': "NÚMERO D'EXPEDIENT:",
            'lic_id': "LICITADOR I IDENTIFICACIÓ DEL RECEPTOR:",
            'empresa': "EMPRESA", 'contacto': "PERSONA DE CONTACTE", 'telefono': "TELÈFON", 'email': "E-MAIL",
            'receptor': "NOM I DNI/MATRÍCULA DEL RECEPTOR", 'fecha': "DATA ENTREGA", 'lugar': "LLOC ENTREGA",
            'id_lic': "IDENTIFICACIÓ ARTICLE LICITAT", 'id_ofr': "IDENTIFICACIÓ ARTICLE DEL LICITADOR",
            'col_lot': "LOT", 'col_art': "ARTICLE", 'col_cod': "CODI HCB",
            'col_tec': "DENOMINACIÓ I REQUERIMENTS TÈCNICS",
            'col_den_lic': "DENOMINACIÓ ARTICLE LICITADOR", 'col_ref': "REFERÈNCIA ARTICLE",
            'col_req': "QUANTITAT DE MOSTRES\nREQUERIDES", 'col_pres': "QUANTITAT DE MOSTRES\nPRESENTADES",
            'sheet': 'Plantilla Annex',
        }
    else:
        t = {
            'title': "ANEXO DE CUMPLIMENTACIÓN OBLIGATORIA 2 PPT ALBARÁN DE MUESTRAS",
            'instr': ("Deberá presentar las muestras relacionadas en este albarán según se indique en los pliegos.\n"
                      "Rellene sólo las celdas en blanco o desprotegidas."),
            'tit_exp': "TÍTULO DEL EXPEDIENTE:", 'num_exp': "NÚMERO DE EXPEDIENTE:",
            'lic_id': "LICITADOR E IDENTIFICACIÓN DEL RECEPTOR:",
            'empresa': "EMPRESA", 'contacto': "PERSONA DE CONTACTO", 'telefono': "TELÉFONO", 'email': "E-MAIL",
            'receptor': "NOMBRE Y DNI/MATRICULA DEL RECEPTOR", 'fecha': "FECHA ENTREGA", 'lugar': "LUGAR ENTREGA",
            'id_lic': "IDENTIFICACIÓN ARTICULO LICITADO", 'id_ofr': "IDENTIFICACIÓN ARTICULO DEL LICITADOR",
            'col_lot': "LOTE", 'col_art': "ARTÍCULO", 'col_cod': "CÓDIGO HCB",
            'col_tec': "DENOMINACIÓN Y REQUISITOS TÉCNICOS",
            'col_den_lic': "DENOMINACIÓN ARTÍCULO LICITADOR", 'col_ref': "REFERENCIA ARTÍCULO",
            'col_req': "CANTIDAD DE MUESTRAS\nREQUERIDAS", 'col_pres': "CANTIDAD DE MUESTRAS\nPRESENTADAS",
            'sheet': 'Plantilla Annex',
        }

    # --- DADES DE PRODUCTES (Full Inici) ---
    ws_in = wb_in['Full Inici']

    mapping, map_warnings = get_column_mapping(wb_in, 'AM', lang=lang)
    if not mapping:
        raise ValueError(f"AM ({lang}): No se ha podido mapear ninguna columna del HI. Revisa el fichero.")
    col_lot = mapping.get(t['col_lot'], mapping.get("LOTE", mapping.get("LOT", 23)))
    col_art = mapping.get(t['col_art'], mapping.get("ARTÍCULO", mapping.get("ARTICLE", 24)))
    col_cod = mapping.get(t['col_cod'], mapping.get("CÓDIGO HCB", mapping.get("CODI HCB", 1)))
    col_tec = mapping.get(t['col_tec'], mapping.get("DENOMINACIÓN Y REQUISITOS TÉCNICOS", mapping.get("DENOMINACIÓ I REQUERIMENTS TÈCNICS", 32)))
    col_req = mapping.get(t['col_req'], mapping.get("CANTIDAD DE MUESTRAS\nREQUERIDAS", mapping.get("QUANTITAT DE MOSTRES\nREQUERIDES", 79)))

    dades_extretes = []
    fila_orig = 6
    while ws_in.cell(row=fila_orig, column=col_art).value is not None or ws_in.cell(row=fila_orig, column=col_cod).value is not None:
        val_art_str = str(ws_in.cell(row=fila_orig, column=col_art).value).upper()
        if "ARTICUL" in val_art_str or "ARTÍCUL" in val_art_str or val_art_str == "ARTICLE":
            fila_orig += 1
            continue

        val_w = ws_in.cell(row=fila_orig, column=col_lot).value
        if val_w is not None and str(val_w).strip() != "" and "LOT" not in str(val_w).upper():
            val_req_val = ws_in.cell(row=fila_orig, column=col_req).value
            try: val_req_val = float(val_req_val) if val_req_val is not None else 1.0
            except: val_req_val = 1.0

            dades_extretes.append({
                "lot": val_w,
                "article": ws_in.cell(row=fila_orig, column=col_art).value,
                "codi_hcb": ws_in.cell(row=fila_orig, column=col_cod).value,
                "tecnic": ws_in.cell(row=fila_orig, column=col_tec).value,
                "requerides": val_req_val
            })
        fila_orig += 1
        if fila_orig > 10000: break

    # --- VALIDACIÓN ---
    issues = list(map_warnings)
    if not dades_extretes:
        issues.insert(0, "No se han extraído filas de datos del HI")
    if issues:
        raise ValueError(f"AM ({lang.upper()}): " + " | ".join(issues))

    cols = [t['col_lot'], t['col_art'], t['col_cod'], t['col_tec'],
            t['col_den_lic'], t['col_ref'], t['col_req'], t['col_pres']]

    df_final = pd.DataFrame(columns=cols)
    df_temp = pd.DataFrame(dades_extretes)
    df_final[t['col_lot']] = df_temp["lot"]
    df_final[t['col_art']] = df_temp["article"]
    df_final[t['col_cod']] = df_temp["codi_hcb"]
    df_final[t['col_tec']] = df_temp["tecnic"]
    df_final[t['col_req']] = df_temp["requerides"]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, header=False, startrow=15, sheet_name=t['sheet'])

    output.seek(0)
    wb_out = load_workbook(output)
    ws = wb_out.active

    # --- ESTILS ---
    align_center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    bold_9 = Font(bold=True, size=9)
    fill_verd = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    fill_gris = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_groc_inst = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    fill_blau = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- MAQUETACIÓ CABECERA ---
    if os.path.exists(logo_path):
        try:
            img = Image(logo_path); img.width = 240; img.height = 100; ws.add_image(img, 'A1')
        except: pass

    ws.row_dimensions[1].height = 85
    apply_style(ws, "C1:H1", t['title'], font=Font(bold=True, size=14))

    apply_style(ws, "A2:H2", t['instr'], font=Font(italic=True, size=10), fill=fill_groc_inst, border=border)
    ws.row_dimensions[2].height = 40

    apply_style(ws, "A4:B4", t['tit_exp'], font=bold_9, fill=fill_gris, border=border)
    apply_style(ws, "C4:H4", titol_expedient, align=align_left_wrap, border=border)

    apply_style(ws, "A5:B5", t['num_exp'], font=bold_9, fill=fill_gris, border=border)
    apply_style(ws, "C5:H5", num_expedient, align=align_left_wrap, border=border)

    apply_style(ws, "A7:H7", t['lic_id'], font=bold_9, fill=fill_gris, border=border)

    apply_style(ws, "A9:B9", t['empresa'], fill=fill_gris, font=bold_9, border=border); apply_style(ws, "C9:D9", border=border)
    apply_style(ws, "A10:B10", t['contacto'], fill=fill_gris, font=bold_9, border=border); apply_style(ws, "C10:D10", border=border)
    apply_style(ws, "A11:B11", t['telefono'], fill=fill_gris, font=bold_9, border=border); apply_style(ws, "C11:D11", border=border)
    apply_style(ws, "A12:B12", t['email'], fill=fill_gris, font=bold_9, border=border); apply_style(ws, "C12:D12", border=border)

    apply_style(ws, "E9:F9", t['receptor'], fill=fill_gris, font=bold_9, align=align_left_wrap, border=border); apply_style(ws, "G9:H9", border=border)
    apply_style(ws, "E10:F10", t['fecha'], fill=fill_gris, font=bold_9, align=align_left_wrap, border=border); apply_style(ws, "G10:H10", border=border)
    apply_style(ws, "E11:F11", t['lugar'], fill=fill_gris, font=bold_9, align=align_left_wrap, border=border); apply_style(ws, "G11:H11", border=border)

    apply_style(ws, "A14:D14", t['id_lic'], fill=fill_verd, font=bold_9, border=border)
    apply_style(ws, "E14:H14", t['id_ofr'], fill=fill_gris, font=bold_9, border=border)

    for c in range(1, 9):
        cell = ws.cell(row=15, column=c)
        cell.font = bold_9; cell.alignment = align_center_wrap; cell.border = border
        if c <= 4 or c == 7: cell.fill = fill_verd
        elif c == 8: cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        else: cell.fill = fill_gris

    ws.cell(row=15, column=1, value=t['col_lot'])
    ws.cell(row=15, column=2, value=t['col_art'])
    ws.cell(row=15, column=3, value=t['col_cod'])
    ws.cell(row=15, column=4, value=t['col_tec'])
    ws.cell(row=15, column=5, value=t['col_den_lic'])
    ws.cell(row=15, column=6, value=t['col_ref'])
    ws.cell(row=15, column=7, value=t['col_req'])
    ws.cell(row=15, column=8, value=t['col_pres'])

    for r in range(4, 13): ws.row_dimensions[r].height = 25
    ws.row_dimensions[14].height = 30
    ws.row_dimensions[15].height = 65

    last_row = ws.max_row
    for r in range(16, last_row + 1):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.alignment = align_center_wrap; cell.border = border
            if c in [1, 2, 3, 4, 7]: cell.protection = Protection(locked=True)
            else: cell.protection = Protection(locked=False)

    for zona in ["C9:D12", "G9:H11"]:
        start, end = zona.split(":")
        for row in ws[start:end]:
            for cell in row: cell.protection = Protection(locked=False)

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    
    ws.protection.password = '1234'
    ws.protection.sheet = True
    
    final_output = io.BytesIO()
    wb_out.save(final_output)
    return final_output.getvalue()

def parse_num(val, default=0.0):
    if val is None: return default
    if isinstance(val, (int, float)): return float(val)
    try: return float(str(val).replace(',', '.').strip())
    except: return default

def generate_oe(input_bytes, logo_path='logo.png', lang='es'):
    wb_in = load_workbook(io.BytesIO(input_bytes), data_only=True)
    nom_pestanya_cab = next((s for s in wb_in.sheetnames if s.upper() == "CABECERA"), None)
    if not nom_pestanya_cab: raise ValueError("No s'ha trobat la pestanya 'cabecera'.")

    ws_cab = wb_in[nom_pestanya_cab]
    titol_expedient = ws_cab['B9'].value
    num_expedient = ws_cab['B5'].value

    ws_in = wb_in['Full Inici']

    # Duración expediente en meses desde pestaña Cabecera (B14)
    duracio_mesos = parse_num(ws_cab['B14'].value, 12.0)

    # --- TEXTOS PER IDIOMA ---
    if lang == 'cat':
        t = {
            'title': "ANNEX DE COMPLIMENTACIÓ OBLIGATÒRIA 3 PCAP OFERTA ECONÒMICA",
            'instr': "Els imports ofertats hauran d'incloure tots els tributs... Només escriviu a les cel·les en blanc.",
            'tit_exp': "TÍTOL DE L'EXPEDIENT:", 'num_exp': "NÚMERO D'EXPEDIENT:",
            'lic_id': "LICITADOR I IDENTIFICACIÓ DE L'OFERTA:",
            'empresa': "EMPRESA", 'email': "E-MAIL", 'contacto': "PERSONA DE CONTACTE", 'telefono': "TELÈFON",
            'fecha': "DATA", 'duracion': "DURACIÓ OFERTA",
            'id_art': "IDENTIFICACIÓ ARTICLE I PREU MÀX. LICITAT", 'oferta_lic': "OFERTA DEL LICITADOR",
            'total': "TOTAL OFERTA ECONÒMICA:",
            'sheet': 'ANNEX OFERTA',
            'col_lot': "LOT", 'col_art': "ARTICLE", 'col_tec': "DENOMINACIÓ I REQUERIMENTS TÈCNICS",
            'col_cod': "CODI HCB", 'col_uml': "UNITAT DE MESURA LICITADA (***)",
            'col_qty': "QUANTITAT",
            'col_pre': "BASE IMPOSABLE MÀXIMA DE LICITACIÓ UNITÀRIA(**) (***)",
            'col_iva': "TIPUS IVA",
            'col_bi_total': "BASE IMPOSABLE MÀXIMA DE LICITACIÓ TOTAL (**)",
            'col_den_lic': "DENOMINACIÓ ARTICLE LICITADOR", 'col_ref_lic': "REFERÈNCIA ARTICLE LICITADOR",
            'col_bi_ofert': "BASE IMPOSABLE UNITÀRIA OFERTADA \n 2 DECIMALS\n (***)",
            'col_iva_ofert': " % IVA OFERTAT",
            'col_bi_total_ofert': "BASE IMPOSABLE TOTAL OFERTADA ",
            'col_import_total': "IMPORT TOTAL amb descompte(IVA INCLÒS)",
        }
    else:
        t = {
            'title': "ANEXO DE CUMPLIMENTACIÓN OBLIGATORIA 3 PCAP OFERTA ECONÓMICA",
            'instr': "Los importes ofertados deberán incluir todos los tributos... Sólo escriba en las celdas en blanco.",
            'tit_exp': "TÍTULO DEL EXPEDIENTE:", 'num_exp': "NÚMERO DE EXPEDIENTE:",
            'lic_id': "LICITADOR E IDENTIFICACIÓN DE LA OFERTA:",
            'empresa': "EMPRESA", 'email': "E-MAIL", 'contacto': "PERSONA DE CONTACTO", 'telefono': "TELÉFONO",
            'fecha': "FECHA", 'duracion': "DURACIÓN OFERTA",
            'id_art': "IDENTIFICACIÓN ARTÍCULO Y PRECIO MÁX. LICITADO", 'oferta_lic': "OFERTA DEL LICITADOR",
            'total': "TOTAL OFERTA ECONÓMICA:",
            'sheet': 'ANNEX OFERTA',
            'col_lot': "LOTE", 'col_art': "ARTÍCULO", 'col_tec': "DENOMINACIÓN Y REQUISITOS TÉCNICOS",
            'col_cod': "CÓDIGO HCB", 'col_uml': "UNIDAD DE MEDIDA LICITADA (***)",
            'col_qty': "CANTIDAD",
            'col_pre': "BASE IMPONIBLE MÁXIMA DE LICITACIÓN UNITARIA(**) (***)",
            'col_iva': "TIPO IVA",
            'col_bi_total': "BASE IMPONIBLE MÁXIMA DE LICITACIÓN TOTAL (**)",
            'col_den_lic': "DENOMINACIÓN ARTÍCULO LICITADOR", 'col_ref_lic': "REFERENCIA ARTÍCULO LICITADOR",
            'col_bi_ofert': "BASE IMPONIBLE UNITARIA OFERTADA \n 2 DECIMALES\n (***)",
            'col_iva_ofert': " % IVA OFERTADO",
            'col_bi_total_ofert': "BASE IMPONIBLE TOTAL OFERTADA ",
            'col_import_total': "IMPORTE TOTAL con descuento(IVA INCLUIDO)",
        }

    mapping, map_warnings = get_column_mapping(wb_in, 'OE', lang=lang)
    if not mapping:
        raise ValueError(f"OE ({lang}): No se ha podido mapear ninguna columna del HI. Revisa el fichero.")
    col_lot = mapping.get(t['col_lot'], mapping.get("LOTE", mapping.get("LOT", 23)))
    col_art = mapping.get(t['col_art'], mapping.get("ARTÍCULO", mapping.get("ARTICLE", 24)))
    col_cod = mapping.get(t['col_cod'], mapping.get("CÓDIGO HCB", mapping.get("CODI HCB", 1)))
    col_tec = mapping.get(t['col_tec'], mapping.get("DENOMINACIÓN Y REQUISITOS TÉCNICOS", mapping.get("DENOMINACIÓ I REQUERIMENTS TÈCNICS", 32)))
    col_uml = mapping.get(t['col_uml'], mapping.get("UNIDAD DE MEDIDA LICITADA (***)", mapping.get("UNITAT DE MESURA LICITADA (***)", 68)))
    col_qty = mapping.get(t['col_qty'], mapping.get("CANTIDAD", mapping.get("QUANTITAT", 69)))
    col_pre = mapping.get(t['col_pre'], mapping.get("BASE IMPONIBLE MÁXIMA DE LICITACIÓN UNITARIA(**) (***)", mapping.get("BASE IMPOSABLE MÀXIMA DE LICITACIÓ UNITÀRIA(**) (***)", 70)))
    col_iva = mapping.get(t['col_iva'], mapping.get("TIPO IVA", mapping.get("TIPUS IVA", 71)))

    dades_extretes = []
    fila_orig = 7
    while ws_in.cell(row=fila_orig, column=col_art).value is not None or ws_in.cell(row=fila_orig, column=col_cod).value is not None:
        val_art_str = str(ws_in.cell(row=fila_orig, column=col_art).value).upper()
        if "ARTICUL" in val_art_str or "ARTÍCUL" in val_art_str or val_art_str == "ARTICLE":
            fila_orig += 1
            continue

        val_w = ws_in.cell(row=fila_orig, column=col_lot).value
        if val_w is not None and str(val_w).strip() != "" and "LOT" not in str(val_w).upper():
            val_qty_val = parse_num(ws_in.cell(row=fila_orig, column=col_qty).value)
            preu_max = parse_num(ws_in.cell(row=fila_orig, column=col_pre).value)
            val_bs = ws_in.cell(row=fila_orig, column=col_iva).value
            try:
                iva = float(val_bs) if isinstance(val_bs, (int, float)) else float(str(val_bs).replace('%', '').replace(',', '.').strip()) / 100.0
            except: iva = 0.0
            dades_extretes.append({
                "lot": val_w, "article": ws_in.cell(row=fila_orig, column=col_art).value,
                "codi_hcb": ws_in.cell(row=fila_orig, column=col_cod).value,
                "tecnic": ws_in.cell(row=fila_orig, column=col_tec).value,
                "uml": ws_in.cell(row=fila_orig, column=col_uml).value,
                "quantitat": (val_qty_val / 12) * duracio_mesos,
                "preu_max": preu_max, "iva": iva
            })
        fila_orig += 1
        if fila_orig > 10000: break

    # --- VALIDACIÓN ---
    issues = list(map_warnings)
    if not dades_extretes:
        issues.insert(0, "No se han extraído filas de datos del HI")
    if issues:
        raise ValueError(f"OE ({lang.upper()}): " + " | ".join(issues))

    cols = [t['col_lot'], t['col_art'], t['col_tec'], t['col_cod'], t['col_uml'], t['col_qty'], t['col_pre'], t['col_iva'], t['col_bi_total'], t['col_den_lic'], t['col_ref_lic'], t['col_bi_ofert'], t['col_iva_ofert'], t['col_bi_total_ofert'], t['col_import_total']]
    df_final = pd.DataFrame(columns=cols)
    df_temp = pd.DataFrame(dades_extretes)
    df_final[t['col_lot']] = df_temp["lot"]; df_final[t['col_art']] = df_temp["article"]; df_final[t['col_tec']] = df_temp["tecnic"]; df_final[t['col_cod']] = df_temp["codi_hcb"]; df_final[t['col_uml']] = df_temp["uml"]; df_final[t['col_qty']] = df_temp["quantitat"]; df_final[t['col_pre']] = df_temp["preu_max"]; df_final[t['col_iva']] = df_temp["iva"]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, header=False, startrow=16, sheet_name=t['sheet'])

    output.seek(0); wb_out = load_workbook(output); ws = wb_out.active

    align_center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    bold_9 = Font(bold=True, size=9); bold_10 = Font(bold=True, size=10)
    fill_verd = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    fill_blau = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    fill_gris = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_taronja = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- CABECERA ---
    if os.path.exists(logo_path):
        try:
            img = Image(logo_path); img.width = 240; img.height = 100; ws.add_image(img, 'A1')
        except: pass

    ws.row_dimensions[1].height = 85
    apply_style(ws, "C1:O1", t['title'], font=Font(bold=True, size=14))

    apply_style(ws, "A2:O2", t['instr'], font=Font(italic=True, size=10), fill=fill_taronja, border=border)
    ws.row_dimensions[2].height = 40

    apply_style(ws, "A4:B4", t['tit_exp'], font=bold_9, fill=fill_gris, border=border)
    apply_style(ws, "C4:O4", titol_expedient, align=align_left_wrap, border=border)

    apply_style(ws, "A5:B5", t['num_exp'], font=bold_9, fill=fill_gris, border=border)
    apply_style(ws, "C5:O5", num_expedient, align=align_left_wrap, border=border)

    apply_style(ws, "A7:O7", t['lic_id'], font=bold_9, fill=fill_gris, border=border)

    # Bloc Licitador
    apply_style(ws, "A9:B9", t['empresa'], fill=fill_gris, font=bold_9, border=border); apply_style(ws, "C9:H9", border=border)
    apply_style(ws, "A10:B10", t['email'], fill=fill_gris, font=bold_9, border=border); apply_style(ws, "C10:H10", border=border)
    apply_style(ws, "A11:B11", t['contacto'], fill=fill_gris, font=bold_9, border=border); apply_style(ws, "C11:H11", border=border)
    apply_style(ws, "A12:B12", t['telefono'], fill=fill_gris, font=bold_9, border=border); apply_style(ws, "C12:H12", border=border)

    apply_style(ws, "I9:K9", t['fecha'], fill=fill_gris, font=bold_9, border=border); apply_style(ws, "L9:O9", border=border)
    apply_style(ws, "I10:K10", t['duracion'], fill=fill_gris, font=bold_9, border=border); apply_style(ws, "L10:O10", border=border)

    apply_style(ws, "A15:I15", t['id_art'], fill=fill_verd, font=bold_9, border=border)
    apply_style(ws, "J15:O15", t['oferta_lic'], fill=fill_taronja, font=bold_9, border=border)
    
    for c_idx, h_text in enumerate(cols, start=1):
        cell = ws.cell(row=16, column=c_idx); cell.value = h_text; cell.font = bold_9; cell.alignment = align_center_wrap; cell.border = border
        if c_idx <= 9: cell.fill = fill_verd
        elif c_idx in [10, 11, 12, 13]: cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"); cell.font = Font(bold=True, size=9, color="FF0000")

    last_row = ws.max_row; format_milers = '#,##0.00'; format_iva = '0%'
    for r in range(17, last_row + 1):
        ws.cell(row=r, column=6).number_format = '#,##0'; ws.cell(row=r, column=7).number_format = format_milers; ws.cell(row=r, column=8).number_format = format_iva
        ws.cell(row=r, column=9).value = f"=F{r}*G{r}"; ws.cell(row=r, column=9).number_format = format_milers
        ws.cell(row=r, column=14).value = f'=IF(L{r}="","",F{r}*L{r})'; ws.cell(row=r, column=14).number_format = format_milers
        ws.cell(row=r, column=15).value = f'=IF(N{r}="","",N{r}*(1+M{r}))'; ws.cell(row=r, column=15).number_format = format_milers
        for c in range(1, 16):
            cell = ws.cell(row=r, column=c)
            if type(cell).__name__ != 'MergedCell':
                cell.alignment = align_center_wrap; cell.border = border
                cell.protection = Protection(locked=c not in [10, 11, 12, 13])

    fill_vermell = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws.conditional_formatting.add(f'N17:O{last_row}', FormulaRule(formula=[f'AND($N17<>"",$N17>$I17)'], fill=fill_vermell))

    row_total = last_row + 1
    for c in range(1, 16): ws.cell(row=row_total, column=c).border = border
    apply_style(ws, f"B{row_total}:F{row_total}", t['total'], font=bold_10, align=Alignment(horizontal='right'), border=border)
    ws.cell(row=row_total, column=9).value = f"=SUM(I17:I{last_row})"; ws.cell(row=row_total, column=9).number_format = format_milers; ws.cell(row=row_total, column=9).font = bold_10
    ws.cell(row=row_total, column=14).value = f"=SUM(N17:N{last_row})"; ws.cell(row=row_total, column=14).number_format = format_milers; ws.cell(row=row_total, column=14).font = bold_10
    ws.cell(row=row_total, column=15).value = f"=SUM(O17:O{last_row})"; ws.cell(row=row_total, column=15).number_format = format_milers; ws.cell(row=row_total, column=15).font = bold_10
    
    # --- ALÇADES I AMPLADES ---
    for r in range(4, 13): ws.row_dimensions[r].height = 25
    ws.row_dimensions[15].height = 40; ws.row_dimensions[16].height = 55; ws.row_dimensions[row_total].height = 30
    
    ws.column_dimensions['A'].width = 8; ws.column_dimensions['B'].width = 10; ws.column_dimensions['C'].width = 45; ws.column_dimensions['D'].width = 12; ws.column_dimensions['E'].width = 12; ws.column_dimensions['F'].width = 12; ws.column_dimensions['G'].width = 15; ws.column_dimensions['H'].width = 12; ws.column_dimensions['I'].width = 15; ws.column_dimensions['J'].width = 35; ws.column_dimensions['K'].width = 15; ws.column_dimensions['L'].width = 15; ws.column_dimensions['M'].width = 12; ws.column_dimensions['N'].width = 15; ws.column_dimensions['O'].width = 15

    # Desbloqueig campos
    for zona in ["C9:H12", "L9:O10"]:
        start, end = zona.split(":")
        for row in ws[start:end]:
            for cell in row: cell.protection = Protection(locked=False)

    ws.protection.password = '1234'; ws.protection.sheet = True
    final_output = io.BytesIO(); wb_out.save(final_output)
    return final_output.getvalue()

def generate_ot(input_bytes, logo_path='logo.png', lang='cat'):
    wb_in = load_workbook(io.BytesIO(input_bytes), data_only=True)
    nom_pestanya_cab = next((s for s in wb_in.sheetnames if s.upper() == "CABECERA"), None)
    if not nom_pestanya_cab: raise ValueError("No s'ha trobat la pestanya 'cabecera'.")

    ws_cab = wb_in[nom_pestanya_cab]
    titol_expedient = ws_cab['B9'].value
    num_expedient = ws_cab['B5'].value

    ws_in = wb_in['Full Inici']

    # Duración expediente en meses desde pestaña Cabecera (B14)
    duracio_mesos = parse_num(ws_cab['B14'].value, 12.0)

    # --- TEXTOS PER IDIOMA ---
    if lang == 'cat':
        t = {
            'title': "ANNEX DE COMPLIMENTACIÓ OBLIGATÒRIA 1 PPT D'OFERTA TÈCNICA (ACO1_PPT_OT)",
            'instr': ("Ompli en aquest annex EN UN ÚNIC FITXER EXCEL, LES OFERTES A TOTS ELS LOTS A QUÈ ES PRESENTI.\n"
                      "En cas que la plataforma electrònica li requereixi pujar un fitxer d'oferta per cada lot..."),
            'tit_exp': "TÍTOL DE L'EXPEDIENT:", 'num_exp': "NÚMERO D'EXPEDIENT:",
            'lic_id': "LICITADOR I IDENTIFICACIÓ DE L'OFERTA:",
            'empresa': "EMPRESA", 'mail': "MAIL", 'contacto': "PERSONA DE CONTACTE (COMERCIAL)",
            'telefono': "TELÈFON", 'mail_contacte': "DIRECCIÓ MAIL CONTACTE (No comandes)",
            'fecha': "DATA", 'mail_cmd': "MAIL comandes", 'tel_cmd': "TELÈFON comandes", 'fax_cmd': "FAX COMANDES",
            'id_lic': "IDENTIFICACIÓ ARTICLE LICITAT", 'id_ofr': "IDENTIFICACIÓ ARTICLE OFERTAT LICITADOR",
            'sheet': 'Annex',
            'col_lot': "LOT", 'col_art': "ARTICLE", 'col_cod': "CODI HCB",
            'col_tec': "DENOMINACIÓ I REQUERIMENTS TÈCNICS",
            'col_qty': "QUANTITAT", 'col_uml': "UNITAT DE MESURA LICITADA (UML)",
            'col_den_lic': "DENOMINACIÓ ARTICLE LICITADOR", 'col_desc_lic': "DESCRIPCIÓ ARTICLE LICITADOR",
            'col_ref_mat': "REFERÈNCIA MATERIAL LICITADOR (***)",
            'col_fmt_pres': "FORMAT DE PRESENTACIÓ", 'col_uml_pres': "UNITATS UML EN PRESENTACIÓ (**)",
            'col_fmt_cmd': "FORMAT PRESENTACIÓ UNITAT MÍNIMA DE COMANDA",
            'col_uml_cmd': "UNITATS UML EN PRESENTACIÓ DE COMANDA (**)",
            'col_fmt_menor': "ALTRE FORMAT DE PRESENTACIÓ MENOR CONTINGUTS (***)",
            'col_uml_menor': "UNITATS UML EN ALTRE FORMAT MENOR (**)",
            'col_fab': "NOM DEL FABRICANT", 'col_marca': "MARCA", 'col_ref_fab': "REF. MATERIAL DEL FABRICANT",
        }
    else:
        t = {
            'title': "ANEXO DE CUMPLIMENTACIÓN OBLIGATORIA 1 PPT OFERTA TÉCNICA (ACO1_PPT_OT)",
            'instr': ("Rellene en este anexo EN UN ÚNICO FICHERO EXCEL, LAS OFERTAS A TODOS LOS LOTES A LOS QUE SE PRESENTE.\n"
                      "En caso de que la plataforma electrónica le requiera subir un fichero de oferta por cada lote..."),
            'tit_exp': "TÍTULO DEL EXPEDIENTE:", 'num_exp': "NÚMERO DE EXPEDIENTE:",
            'lic_id': "LICITADOR E IDENTIFICACIÓN DE LA OFERTA:",
            'empresa': "EMPRESA", 'mail': "E-MAIL", 'contacto': "PERSONA DE CONTACTO (COMERCIAL)",
            'telefono': "TELÉFONO", 'mail_contacte': "DIRECCIÓN E-MAIL CONTACTO (No pedidos)",
            'fecha': "FECHA", 'mail_cmd': "E-MAIL pedidos", 'tel_cmd': "TELÉFONO pedidos", 'fax_cmd': "FAX PEDIDOS",
            'id_lic': "IDENTIFICACIÓN ARTÍCULO LICITADO", 'id_ofr': "IDENTIFICACIÓN ARTÍCULO OFERTADO LICITADOR",
            'sheet': 'Annex',
            'col_lot': "LOTE", 'col_art': "ARTÍCULO", 'col_cod': "CÓDIGO HCB",
            'col_tec': "DENOMINACIÓN Y REQUISITOS TÉCNICOS",
            'col_qty': "CANTIDAD", 'col_uml': "UNIDAD DE MEDIDA LICITADA (UML)",
            'col_den_lic': "DENOMINACIÓN ARTÍCULO LICITADOR", 'col_desc_lic': "DESCRIPCIÓN ARTÍCULO LICITADOR",
            'col_ref_mat': "REFERENCIA MATERIAL LICITADOR (***)",
            'col_fmt_pres': "FORMATO DE PRESENTACIÓN", 'col_uml_pres': "UNIDADES UML EN PRESENTACIÓN (**)",
            'col_fmt_cmd': "FORMATO PRESENTACIÓN UNIDAD MÍNIMA DE PEDIDO",
            'col_uml_cmd': "UNIDADES UML EN PRESENTACIÓN DE PEDIDO (**)",
            'col_fmt_menor': "OTRO FORMATO DE PRESENTACIÓN MENOR CONTENIDOS (***)",
            'col_uml_menor': "UNIDADES UML EN OTRO FORMATO MENOR (**)",
            'col_fab': "NOMBRE DEL FABRICANTE", 'col_marca': "MARCA", 'col_ref_fab': "REF. MATERIAL DEL FABRICANTE",
        }

    mapping, map_warnings = get_column_mapping(wb_in, 'OT', lang=lang)
    if not mapping:
        raise ValueError(f"OT ({lang}): No se ha podido mapear ninguna columna del HI. Revisa el fichero.")
    col_lot = mapping.get(t['col_lot'], mapping.get("LOT", mapping.get("LOTE", 23)))
    col_art = mapping.get(t['col_art'], mapping.get("ARTICLE", mapping.get("ARTÍCULO", 24)))
    col_cod = mapping.get(t['col_cod'], mapping.get("CODI HCB", mapping.get("CÓDIGO HCB", 1)))
    col_tec = mapping.get(t['col_tec'], mapping.get("DENOMINACIÓ I REQUERIMENTS TÈCNICS", mapping.get("DENOMINACIÓN Y REQUISITOS TÉCNICOS", 32)))
    col_uml = mapping.get(t['col_uml'], mapping.get("UNITAT DE MESURA LICITADA (UML)", mapping.get("UNIDAD DE MEDIDA LICITADA (UML)", 68)))
    col_qty = mapping.get(t['col_qty'], mapping.get("QUANTITAT", mapping.get("CANTIDAD", 69)))

    dades_extretes = []
    fila_orig = 6
    while ws_in.cell(row=fila_orig, column=col_art).value is not None or ws_in.cell(row=fila_orig, column=col_cod).value is not None:
        val_art_str = str(ws_in.cell(row=fila_orig, column=col_art).value).upper()
        if "ARTICUL" in val_art_str or "ARTÍCUL" in val_art_str or val_art_str == "ARTICLE":
            fila_orig += 1
            continue

        val_w = ws_in.cell(row=fila_orig, column=col_lot).value
        if val_w is not None and str(val_w).strip() != "" and "LOT" not in str(val_w).upper():
            val_qty_val = parse_num(ws_in.cell(row=fila_orig, column=col_qty).value)
            dades_extretes.append({
                "lot": val_w, "article": ws_in.cell(row=fila_orig, column=col_art).value,
                "codi_hcb": ws_in.cell(row=fila_orig, column=col_cod).value,
                "tecnic": ws_in.cell(row=fila_orig, column=col_tec).value,
                "uml": ws_in.cell(row=fila_orig, column=col_uml).value, "quantitat": (val_qty_val / 12) * duracio_mesos
            })
        fila_orig += 1
        if fila_orig > 10000: break

    # --- VALIDACIÓN ---
    issues = list(map_warnings)
    if not dades_extretes:
        issues.insert(0, "No se han extraído filas de datos del HI")
    if issues:
        raise ValueError(f"OT ({lang.upper()}): " + " | ".join(issues))

    cols = [t['col_lot'], t['col_art'], t['col_cod'], t['col_tec'],
            t['col_qty'], t['col_uml'], t['col_den_lic'],
            t['col_desc_lic'], t['col_ref_mat'],
            t['col_fmt_pres'], t['col_uml_pres'],
            t['col_fmt_cmd'], t['col_uml_cmd'],
            t['col_fmt_menor'], t['col_uml_menor'],
            t['col_fab'], t['col_marca'], t['col_ref_fab']]

    df_final = pd.DataFrame(columns=cols)
    df_temp = pd.DataFrame(dades_extretes)
    df_final[t['col_lot']] = df_temp["lot"]
    df_final[t['col_art']] = df_temp["article"]
    df_final[t['col_cod']] = df_temp["codi_hcb"]
    df_final[t['col_tec']] = df_temp["tecnic"]
    df_final[t['col_qty']] = df_temp["quantitat"]
    df_final[t['col_uml']] = df_temp["uml"]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, startrow=15, sheet_name=t['sheet'])

    output.seek(0); wb_out = load_workbook(output); ws = wb_out.active

    # --- ESTILS ---
    align_center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    bold_9 = Font(bold=True, size=9)
    fill_verd = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    fill_gris = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_groc_inst = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- MAQUETACIÓ CABECERA ---
    if os.path.exists(logo_path):
        try:
            img = Image(logo_path); img.width = 240; img.height = 100; ws.add_image(img, 'A1')
        except: pass

    ws.row_dimensions[1].height = 85
    apply_style(ws, "C1:R1", t['title'], font=Font(bold=True, size=14))

    apply_style(ws, "A2:R2", t['instr'], font=Font(italic=True, size=10), fill=fill_groc_inst, border=border)
    ws.row_dimensions[2].height = 65

    apply_style(ws, "A4:B4", t['tit_exp'], font=bold_9, fill=fill_gris, border=border)
    apply_style(ws, "C4:R4", titol_expedient, align=align_left_wrap, border=border)

    apply_style(ws, "A5:B5", t['num_exp'], font=bold_9, fill=fill_gris, border=border)
    apply_style(ws, "C5:R5", num_expedient, align=align_left_wrap, border=border)

    apply_style(ws, "A7:R7", t['lic_id'], font=bold_9, fill=fill_gris, border=border)

    # Bloc Esquerre Licitador
    apply_style(ws, "A9:B9", t['empresa'], fill=fill_gris, font=bold_9, border=border); apply_style(ws, "C9:I9", border=border)
    apply_style(ws, "A10:B10", t['mail'], fill=fill_gris, font=bold_9, border=border); apply_style(ws, "C10:I10", border=border)
    apply_style(ws, "A11:B11", t['contacto'], fill=fill_gris, font=bold_9, border=border); apply_style(ws, "C11:I11", border=border)
    apply_style(ws, "A12:B12", t['telefono'], fill=fill_gris, font=bold_9, border=border); apply_style(ws, "C12:I12", border=border)
    apply_style(ws, "A13:B13", t['mail_contacte'], fill=fill_gris, font=bold_9, align=align_left_wrap, border=border); apply_style(ws, "C13:I13", border=border)

    # Bloc Dreta Licitador
    apply_style(ws, "J9:L9", t['fecha'], fill=fill_gris, font=bold_9, border=border); apply_style(ws, "M9:N9", border=border)
    apply_style(ws, "J10:L10", t['mail_cmd'], fill=fill_gris, font=bold_9, border=border); apply_style(ws, "M10:N10", border=border)
    apply_style(ws, "J11:L11", t['tel_cmd'], fill=fill_gris, font=bold_9, border=border); apply_style(ws, "M11:N11", border=border)
    apply_style(ws, "J12:L12", t['fax_cmd'], fill=fill_gris, font=bold_9, border=border); apply_style(ws, "M12:N12", border=border)
    apply_style(ws, "O9:R13", "", fill=PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"), border=border)

    # Capçaleres Taula
    apply_style(ws, "A15:F15", t['id_lic'], fill=fill_verd, font=bold_9, border=border)
    apply_style(ws, "G15:R15", t['id_ofr'], fill=fill_gris, font=bold_9, border=border)
    for c in range(1, 19):
        cell = ws.cell(row=16, column=c)
        cell.font = bold_9; cell.alignment = align_center_wrap; cell.border = border
        if c <= 4 or c == 6: cell.fill = fill_verd
        elif c == 5: cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        else: cell.fill = fill_gris

    # --- PROTECCIÓ I ALÇADES ---
    for r in range(4, 14): ws.row_dimensions[r].height = 25
    ws.row_dimensions[16].height = 45

    last_row = ws.max_row
    for r in range(17, last_row + 1):
        ws.row_dimensions[r].height = None 
        val_cant = ws.cell(row=r, column=5).value
        # Si val_cant es 0, bloquegem tota la fila. Si no, bloquegem columnes de HCB (1-6)
        is_zero = False
        try: is_zero = float(val_cant) == 0
        except: pass
        
        for c in range(1, 19):
            cell = ws.cell(row=r, column=c)
            cell.alignment = align_center_wrap; cell.border = border
            if c == 5: cell.number_format = '#,##0'
            cell.protection = Protection(locked=True if is_zero else (c < 7))

    # Desbloqueig camps d'entrada superiors
    for zona in ["C9:I13", "M9:N12", "O9:R13"]:
        start, end = zona.split(":")
        for row in ws[start:end]:
            for cell in row: cell.protection = Protection(locked=False)

    ws.column_dimensions['D'].width = 55
    ws.protection.password = '1234'
    ws.protection.sheet = True
    
    final_output = io.BytesIO()
    wb_out.save(final_output)
    return final_output.getvalue()
