import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Protection, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.drawing.image import Image

# 1. Configuración
fitxer_origen = 'hi.xlsm'
nom_sortida = 'ACO3_PCAP_OE.xlsx'
password_excel = '1234'

# --- DINAMIC COLUMN MAPPING ---
def get_column_mapping(wb_in, annex_type):
    try:
        cab_file = 'CABECERAS.xlsx'
        df_map = pd.read_excel(cab_file, sheet_name=annex_type, header=None)
        annex_headers = [str(h).strip() for h in df_map.iloc[0].fillna("").tolist()]
        hi_headers_to_find = [str(h).strip().upper() for h in df_map.iloc[1].fillna("").tolist()]
        
        ws_in = wb_in['Full Inici']
        mapping = {}
        
        # Scan headers in rows 5 and 6
        for r_idx in [5, 6]:
            row_vals = [str(ws_in.cell(row=r_idx, column=c).value).strip().upper() for c in range(1, 150)]
            for ah, htf in zip(annex_headers, hi_headers_to_find):
                if htf == "NAN" or htf == "": continue
                if htf in row_vals:
                    mapping[ah] = row_vals.index(htf) + 1 # 1-based column index
        return mapping
    except Exception as e:
        print(f"Error loading mapping: {e}")
        return {}

try:
    print(f"Llegint arxiu d'origen: {fitxer_origen}...")
    # data_only=True per llegir valors calculats de les fórmules
    wb_in = load_workbook(fitxer_origen, data_only=True)
    
    # --- DADES DE CABECERA ---
    nom_pestanya_cab = next((s for s in wb_in.sheetnames if s.upper() == "CABECERA"), None)
    if not nom_pestanya_cab:
        raise ValueError("No s'ha trobat la pestanya 'cabecera' al fitxer original.")
    
    ws_cab = wb_in[nom_pestanya_cab]
    titol_expedient = ws_cab['B9'].value
    num_expedient = ws_cab['B5'].value

    # --- DADES DE PRODUCTES (Full Inici) ---
    ws_in = wb_in['Full Inici']
    val_dh3 = ws_in['DH3'].value
    factor_dh3 = float(val_dh3) if val_dh3 is not None else 1.0
    
    mapping = get_column_mapping(wb_in, 'OE')
    print(f"Mapeig detectat: {mapping}")

    col_lot = mapping.get("LOTE", 23)
    col_art = mapping.get("ARTÍCULO", 24)
    col_cod = mapping.get("CÓDIGO HCB", 1)
    col_tec = mapping.get("DENOMINACIÓN Y REQUISITOS TÉCNICOS", 32)
    col_uml = mapping.get("UNIDAD DE MEDIDA LICITADA (***)", 68)
    col_qty = mapping.get("CANTIDAD", 69)
    col_pre = mapping.get("BASE IMPONIBLE MÁXIMA DE LICITACIÓN UNITARIA(**) (***)", 70)
    col_iva = mapping.get("TIPO IVA", 71)

    # Funció auxiliar per parsejar valors numèrics amb format europeu (coma decimal)
    def parse_num(val, default=0.0):
        if val is None:
            return default
        if isinstance(val, (int, float)):
            return float(val)
        try:
            return float(str(val).replace(',', '.').strip())
        except:
            return default

    dades_extretes = []
    fila_orig = 7
    while ws_in.cell(row=fila_orig, column=col_art).value is not None or ws_in.cell(row=fila_orig, column=col_cod).value is not None:
        val_w = ws_in.cell(row=fila_orig, column=col_lot).value 
        if val_w is not None and str(val_w).strip() != "" and str(val_w).upper() != "NUMERO":
            val_qty_val = parse_num(ws_in.cell(row=fila_orig, column=col_qty).value)
            preu_max = parse_num(ws_in.cell(row=fila_orig, column=col_pre).value)
            val_bs = ws_in.cell(row=fila_orig, column=col_iva).value
            try: 
                if isinstance(val_bs, (int, float)):
                    iva = float(val_bs)
                else:
                    iva_val = str(val_bs).replace('%', '').replace(',', '.').strip()
                    iva = float(iva_val) / 100.0 if iva_val else 0.0
            except: 
                iva = 0.0

            quantitat_calculada = (val_qty_val / 12) * factor_dh3

            dades_extretes.append({
                "lot": val_w, 
                "article": ws_in.cell(row=fila_orig, column=col_art).value,
                "codi_hcb": ws_in.cell(row=fila_orig, column=col_cod).value, 
                "tecnic": ws_in.cell(row=fila_orig, column=col_tec).value,
                "uml": ws_in.cell(row=fila_orig, column=col_uml).value,
                "quantitat": quantitat_calculada,
                "preu_max": preu_max,
                "iva": iva
            })
        fila_orig += 1
        if fila_orig > 10000: break

    # --- CREACIÓ DEL DATAFRAME ---
    # 15 Columnes de la plantilla ACO3
    cols = [
        "LOTE", "ARTÍCULO", "DENOMINACIÓN Y REQUISITOS TÉCNICOS",
        "CÓDIGO HCB", "UNIDAD DE MEDIDA LICITADA (***)", "CANTIDAD",
        "BASE IMPONIBLE MÁXIMA DE LICITACIÓN UNITARIA(**) (***)", "TIPO IVA", 
        "BASE IMPONIBLE MÁXIMA DE LICITACIÓN TOTAL (**)",
        "DENOMINACIÓN ARTÍCULO LICITADOR", "REFERENCIA ARTÍCULO LICITADOR", 
        "BASE IMPONIBLE UNITARIA OFERTADA \n 2 DECIMALES\n (***)",
        " % IVA OFERTADO", "BASE IMPONIBLE TOTAL OFERTADA ", 
        "IMPORTE TOTAL con descuento(IVA INCLUIDO)"
    ]

    df_final = pd.DataFrame(columns=cols)
    df_temp = pd.DataFrame(dades_extretes)
    
    df_final["LOTE"] = df_temp["lot"]
    df_final["ARTÍCULO"] = df_temp["article"]
    df_final["DENOMINACIÓN Y REQUISITOS TÉCNICOS"] = df_temp["tecnic"]
    df_final["CÓDIGO HCB"] = df_temp["codi_hcb"]
    df_final["UNIDAD DE MEDIDA LICITADA (***)"] = df_temp["uml"]
    df_final["CANTIDAD"] = df_temp["quantitat"]
    df_final["BASE IMPONIBLE MÁXIMA DE LICITACIÓN UNITARIA(**) (***)"] = df_temp["preu_max"]
    df_final["TIPO IVA"] = df_temp["iva"]
    
    # La columna 9 es fórmula: =F(Cant) * G(BI Max Unitaria)
    # L'escriurem dinàmicament en OpenPyXL o aquí pre-calculat (millor a OpenPyXL per mantenir plantilles dinàmiques).
    # Per Pandas li posarem string buit i ho reomplim a OpenPyxl.
    
    with pd.ExcelWriter(nom_sortida, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, header=False, startrow=16, sheet_name='ANNEX OFERTA')

    wb_out = load_workbook(nom_sortida)
    ws = wb_out.active

    # --- ESTILS ---
    align_center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    bold_9 = Font(bold=True, size=9)
    bold_10 = Font(bold=True, size=10)
    
    fill_verd = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    fill_blau = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    fill_gris = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_groc = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_taronja = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def apply_style(rng, text=None, fill=None, font=None, align=align_center_wrap):
        start_coords, end_coords = rng.split(":") if ":" in rng else (rng, rng)
        if ":" in rng: ws.merge_cells(rng)
        
        start_cell = ws[start_coords]
        if text is not None: start_cell.value = text
        
        # Aplicar estil a TOTA la zona (imprescindible per als bordes en cel·les combinades)
        for r_idx in range(ws[start_coords].row, ws[end_coords].row + 1):
            for c_idx in range(ws[start_coords].column, ws[end_coords].column + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = border
                cell.alignment = align
                if fill: cell.fill = fill
                if font: cell.font = font

    # --- MAQUETACIÓ CABECERA ---
    # Inserció de Logo
    try:
        img = Image('logo.png')
        # Ajustar el logo (proporcionalment sol ser millor mantenir l'aspecte)
        img.width = 240
        img.height = 100
        ws.add_image(img, 'A1')
    except Exception as img_err:
        print(f"Avís: No s'ha pogut carregar el logo: {img_err}")

    ws.row_dimensions[1].height = 85 # Ajustar alçada per al logo i títol
    apply_style("C1:O1", "ANEXO DE CUMPLIMENTACIÓN OBLIGATORIA 3 PCAP OFERTA ECONÓMICA", font=Font(bold=True, size=14), fill=fill_blau)
    
    txt_instr = ("Los importes ofertados deberán incluir todos los tributos, tasas y cánones de cualquier índole, así como cualquier gasto que se ocasione de la ejecución del suministro. IMPUESTOS APLICABLES EXCLUIDOS.\n"
                 "Sólo escriba en las celdas en blanco o desprotegidas.")
    apply_style("A2:O2", txt_instr, font=Font(italic=True, size=10), fill=fill_taronja)
    ws.row_dimensions[2].height = 40

    apply_style("A4:B4", "TÍTULO DEL EXPEDIENTE:", font=bold_9, fill=fill_gris)
    apply_style("C4:O4", titol_expedient, align=align_left_wrap)
    
    apply_style("A5:B5", "NÚMERO DE EXPEDIENTE:", font=bold_9, fill=fill_gris)
    apply_style("C5:O5", num_expedient, align=align_left_wrap)

    apply_style("A7:O7", "LICITADOR E IDENTIFICACIÓN DE LA OFERTA:", font=bold_9, fill=fill_gris)

    # Bloc Esquerre Licitador
    apply_style("A9:B9", "EMPRESA", fill=fill_gris, font=bold_9); apply_style("C9:H9")
    apply_style("A10:B10", "E-MAIL", fill=fill_gris, font=bold_9); apply_style("C10:H10")
    apply_style("A11:B11", "PERSONA DE CONTACTO (COMERCIAL)", fill=fill_gris, font=bold_9); apply_style("C11:H11")
    apply_style("A12:B12", "TELÉFONO", fill=fill_gris, font=bold_9); apply_style("C12:H12")

    # Bloc Dreta Identificacio Oferta
    apply_style("I9:K9", "FECHA", fill=fill_gris, font=bold_9); apply_style("L9:O9")
    apply_style("I10:K10", "DURACIÓN DE LA OFERTA", fill=fill_gris, font=bold_9); apply_style("L10:O10")
    
    # Capçaleres Taula (Fila 15 en la plantilla original)
    apply_style("A15:I15", "IDENTIFICACIÓN ARTÍCULO Y PRECIO MÁX.  LICITADO\n(A rellenar por el órgano de Contratación)", fill=fill_verd, font=bold_9)
    apply_style("J15:O15", "OFERTA DEL LICITADOR\n(A rellenar por el licitador)", fill=fill_taronja, font=bold_9)
    
    headers_taula = cols
    for c_idx, h_text in enumerate(headers_taula, start=1):
        cell = ws.cell(row=16, column=c_idx)
        cell.value = h_text
        cell.font = bold_9
        cell.alignment = align_center_wrap
        cell.border = border
        
        if c_idx <= 9:
            cell.fill = fill_verd
        elif c_idx in [10, 11, 12, 13]:
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # Blanc per editable
            # Posar text vermell i negreta a les capçaleres editables
            cell.font = Font(bold=True, size=9, color="FF0000") 

    # --- INSERCIÓ DE FÓRMULES I FORMAT DE DADES ---
    last_row = ws.max_row
    # format_milers con # al final a veces deja una coma residual en formatos locales. 
    # Usaremos 'General' para la columna de entrada y un formato más limpio para el resto.
    format_milers = '#,##0.00'  # Formato financiero estándar con 2 decimales para evitar la coma residual
    format_iva = '0%'

    for r in range(17, last_row + 1):
        ws.row_dimensions[r].height = None 
        
        # Col 6 - Cantidad
        ws.cell(row=r, column=6).number_format = '#,##0'
        # Col 7 - BI Unitaria Max (G)
        ws.cell(row=r, column=7).number_format = format_milers
        # Col 8 - Tipo IVA Max (H)
        ws.cell(row=r, column=8).number_format = format_iva
        
        # Col 9 - BI MAX LICITACION TOTAL (I) = Cantidad(F) * BI Max Unitaria(G)
        ws.cell(row=r, column=9).value = f"=F{r}*G{r}"
        ws.cell(row=r, column=9).number_format = format_milers

        # Col 12 - BI Unitaria OFERTADA (L) - Usamos General para que no ponga la coma en enteros
        ws.cell(row=r, column=12).number_format = 'General'
        
        # Col 13 - % IVA Ofertado (M)
        ws.cell(row=r, column=13).number_format = format_iva
        
        # Col 14 - BI TOTAL OFERTADA (N) = Cantidad(F) * BI Unitaria Ofertada(L)
        cel_14 = ws.cell(row=r, column=14)
        cel_14.value = f'=IF(L{r}="","",F{r}*L{r})'
        cel_14.number_format = format_milers
        
        # Col 15 - IMPORTE TOTAL IVA INCLUIDO (O) = BI TOTAL OFERTADA(N) + (BI TOTAL OFERTADA(N) * % IVA OFERTADO(M))
        cel_15 = ws.cell(row=r, column=15)
        cel_15.value = f'=IF(N{r}="","",N{r}*(1+M{r}))'
        cel_15.number_format = format_milers

        for c in range(1, 16):
            cell = ws.cell(row=r, column=c)
            if type(cell).__name__ != 'MergedCell':
                cell.alignment = align_center_wrap
                cell.border = border
                
                # Bloquejar tot excepte les columnes 10 a 13 (J-M)
                if c in [10, 11, 12, 13]: 
                    cell.protection = Protection(locked=False)
                else:
                    cell.protection = Protection(locked=True)


    # --- FORMAT CONDICIONAL: BI Total Ofertada (N) > BI Max Total (I) => fons VERMELL en columnes N i O ---
    fill_vermell = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws.conditional_formatting.add(
        f'N17:O{last_row}',
        FormulaRule(formula=[f'AND($N17<>"",$N17>$I17)'], fill=fill_vermell)
    )

    # --- FILA DE TOTALS ---
    row_total = last_row + 1
    
    # Buit de A a N (14 posicions, per la C_row merge)
    for c in range(1, 15): # A-N
        cell_t = ws.cell(row=row_total, column=c)
        cell_t.border = border
    
    cell_label = ws.cell(row=row_total, column=2) # B
    cell_label.value = "TOTAL OFERTA ECONÓMICA:"
    cell_label.font = bold_10
    cell_label.alignment = Alignment(horizontal='right', vertical='center')
    ws.merge_cells(f"B{row_total}:F{row_total}") # B-F
    
    # Total BI MAX Licitacion (Col 9) (I)
    cel_t_9 = ws.cell(row=row_total, column=9)
    cel_t_9.value = f"=SUM(I17:I{last_row})"
    cel_t_9.number_format = format_milers
    cel_t_9.font = bold_10
    cel_t_9.border = border
    
    # Total BI Ofertada (Col 14) (N)
    cel_t_14 = ws.cell(row=row_total, column=14)
    cel_t_14.value = f"=SUM(N17:N{last_row})"
    cel_t_14.number_format = format_milers
    cel_t_14.font = bold_10
    cel_t_14.border = border

    # Total Importe IVA incluido (Col 15) (O)
    cel_t_15 = ws.cell(row=row_total, column=15)
    cel_t_15.value = f"=SUM(O17:O{last_row})"
    cel_t_15.number_format = format_milers
    cel_t_15.font = bold_10
    cel_t_15.border = border
    
    # Res protegit/bloquejat per defecte a la fila totals
    for c in range(1, 16): # A-O
        cell_t = ws.cell(row=row_total, column=c)
        if type(cell_t).__name__ != 'MergedCell':
            cell_t.protection = Protection(locked=True)

    # --- AJUSTOS ALÇADES I AMPLADA DE COLUMNES ---
    for r in range(4, 13): ws.row_dimensions[r].height = 25
    ws.row_dimensions[15].height = 40
    ws.row_dimensions[16].height = 55
    ws.row_dimensions[row_total].height = 30

    # Desbloqueig camps d'entrada superiors (Licitador i dades oferta)
    for zona in ["C9:H12", "L9:O10"]:
        start, end = zona.split(":")
        for row in ws[start:end]:
            for cell in row: cell.protection = Protection(locked=False)

    ws.column_dimensions['A'].width = 8  # Lote
    ws.column_dimensions['B'].width = 10 # Articulo
    ws.column_dimensions['C'].width = 45 # Denominacion tecnica
    ws.column_dimensions['D'].width = 12 # Codigo HCB
    ws.column_dimensions['E'].width = 12 # UML
    ws.column_dimensions['F'].width = 12 # Cantidad
    ws.column_dimensions['G'].width = 15 # Precio Unit Max
    ws.column_dimensions['H'].width = 12 # Tipo IVA
    ws.column_dimensions['I'].width = 15 # BI Max Total
    ws.column_dimensions['J'].width = 35 # Denom Licitador
    ws.column_dimensions['K'].width = 15 # Ref Licitador
    ws.column_dimensions['L'].width = 15 # BI Ofertada
    ws.column_dimensions['M'].width = 12 # % IVA
    ws.column_dimensions['N'].width = 15 # BI Total Ofertada
    ws.column_dimensions['O'].width = 15 # Importe Total
    
    ws.protection.password = password_excel
    ws.protection.sheet = True
    
    wb_out.save(nom_sortida)
    print(f"[OK] Generat amb exit: {nom_sortida}")

except Exception as e:
    import traceback
    print("Error en l'execució:")
    traceback.print_exc()
