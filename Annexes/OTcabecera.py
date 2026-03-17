import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Protection, Border, Side

# 1. Configuración
fitxer_origen = 'hi.xlsm'
nom_sortida = 'ACO1_PPT_OT.xlsx'
password_excel = '1234'

try:
    print("Llegint arxiu d'origen: hi.xlsm...")
    # data_only=True per llegir valors calculats de les fórmules
    wb_in = load_workbook(fitxer_origen, data_only=True)
    
    # --- DADES DE CABECERA ---
    nom_pestanya_cab = next((s for s in wb_in.sheetnames if s.upper() == "CABECERA"), None)
    if not nom_pestanya_cab:
        raise ValueError("No s'ha trobat la pestanya 'cabecera' al fitxer original.")
    
    ws_cab = wb_in[nom_pestanya_cab]
    titol_expedient = ws_cab['B9'].value
    num_expedient = ws_cab['B5'].value

    # --- DADES DE PRODUCTES (Full Inici) ---
    ws_in = wb_in['Full Inici']
    val_dh3 = ws_in['DH3'].value
    factor_dh3 = float(val_dh3) if val_dh3 is not None else 1.0
    
    dades_extretes = []
    fila_orig = 6
    while ws_in[f'X{fila_orig}'].value is not None or ws_in[f'A{fila_orig}'].value is not None:
        val_w = ws_in[f'W{fila_orig}'].value 
        if val_w is not None and str(val_w).strip() != "":
            val_bq = ws_in[f'BQ{fila_orig}'].value
            try: val_bq = float(val_bq) if val_bq is not None else 0.0
            except: val_bq = 0.0
            dades_extretes.append({
                "lot": val_w, "article": ws_in[f'X{fila_orig}'].value,
                "codi_hcb": ws_in[f'A{fila_orig}'].value, "tecnic": ws_in[f'AF{fila_orig}'].value,
                "uml": ws_in[f'BP{fila_orig}'].value, "quantitat": (val_bq / 12) * factor_dh3
            })
        fila_orig += 1
        if fila_orig > 10000: break

    # --- CREACIÓ DEL DATAFRAME ---
    cols = ["LOT", "ARTICLE", "CODI HCB", "DENOMINACIÓ I REQUERIMENTS TÈCNICS", 
            "QUANTITAT", "UNITAT DE MESURA LICITADA (UML)", "DENOMINACIÓ ARTICLE LICITADOR", 
            "DESCRIPCIÓ ARTICLE LICITADOR", "REFERÈNCIA MATERIAL LICITADOR (***)", 
            "FORMAT DE PRESENTACIÓ", "UNITATS UML EN PRESENTACIÓ (**)", 
            "FORMAT PRESENTACIÓ UNITAT MÍNIMA DE COMANDA", "UNITATS UML EN PRESENTACIÓ DE COMANDA (**)", 
            "ALTRE FORMAT DE PRESENTACIÓ MENOR CONTINGUTS (***)", "UNITATS UML EN ALTRE FORMAT MENOR (**)", 
            "NOM DEL FABRICANT", "MARCA", "REF. MATERIAL DEL FABRICANT"]

    df_final = pd.DataFrame(columns=cols)
    df_temp = pd.DataFrame(dades_extretes)
    df_final["LOT"] = df_temp["lot"]
    df_final["ARTICLE"] = df_temp["article"]
    df_final["CODI HCB"] = df_temp["codi_hcb"]
    df_final["DENOMINACIÓ I REQUERIMENTS TÈCNICS"] = df_temp["tecnic"]
    df_final["QUANTITAT"] = df_temp["quantitat"]
    df_final["UNITAT DE MESURA LICITADA (UML)"] = df_temp["uml"]

    with pd.ExcelWriter(nom_sortida, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, startrow=15, sheet_name='Annex')

    wb_out = load_workbook(nom_sortida)
    ws = wb_out.active

    # --- ESTILS ---
    align_center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    bold_9 = Font(bold=True, size=9)
    fill_verd = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    fill_gris = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_groc_inst = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def apply_style(rng, text=None, fill=None, font=None, align=align_center_wrap):
        if ":" in rng: ws.merge_cells(rng)
        start_cell = ws[rng.split(":")[0]]
        if text is not None: start_cell.value = text
        start_coords, end_coords = rng.split(":") if ":" in rng else (rng, rng)
        for row in ws[start_coords:end_coords]:
            for cell in row:
                cell.alignment = align
                cell.border = border
                if fill: cell.fill = fill
                if font: cell.font = font

    # --- MAQUETACIÓ CABECERA ---
    apply_style("A1:R1", "ANNEX DE COMPLIMENTACIÓ OBLIGATÒRIA 1 PPT D'OFERTA TÈCNICA (ACO1_PPT_OT)", font=Font(bold=True, size=14))
    
    txt_instr = ("Ompli en aquest annex EN UN ÚNIC FITXER EXCEL, LES OFERTES A TOTS ELS LOTS A QUÈ ES PRESENTI.\n"
                 "En cas que la plataforma electrònica li requereixi pujar un fitxer d'oferta per cada lot...")
    apply_style("A2:R2", txt_instr, font=Font(italic=True, size=10), fill=fill_groc_inst)
    ws.row_dimensions[2].height = 65

    apply_style("A4:B4", "TÍTOL DE L´EXPEDIENT:", font=bold_9, fill=fill_gris)
    apply_style("C4:R4", titol_expedient, align=align_left_wrap)
    
    apply_style("A5:B5", "NÚMERO D´EXPEDIENT:", font=bold_9, fill=fill_gris)
    apply_style("C5:R5", num_expedient, align=align_left_wrap)

    apply_style("A7:R7", "LICITADOR I IDENTIFICACIÓ DE L'OFERTA:", font=bold_9, fill=fill_gris)

    # Bloc Esquerre Licitador
    apply_style("A9:B9", "EMPRESA", fill=fill_gris, font=bold_9); apply_style("C9:I9")
    apply_style("A10:B10", "MAIL", fill=fill_gris, font=bold_9); apply_style("C10:I10")
    apply_style("A11:B11", "PERSONA DE CONTACTE (COMERCIAL)", fill=fill_gris, font=bold_9); apply_style("C11:I11")
    apply_style("A12:B12", "TELÈFON", fill=fill_gris, font=bold_9); apply_style("C12:I12")
    apply_style("A13:B13", "DIRECCIÓ MAIL CONTACTE (No comandes)", fill=fill_gris, font=bold_9, align=align_left_wrap); apply_style("C13:I13")

    # Bloc Dreta Licitador (CORREGIT)
    apply_style("J9:L9", "DATA", fill=fill_gris, font=bold_9); apply_style("M9:N9")
    apply_style("J10:L10", "MAIL comandes", fill=fill_gris, font=bold_9); apply_style("M10:N10")
    apply_style("J11:L11", "TELÈFON comandes", fill=fill_gris, font=bold_9); apply_style("M11:N11")
    apply_style("J12:L12", "FAX COMANDES", fill=fill_gris, font=bold_9); apply_style("M12:N12")
    apply_style("O9:R13", "", fill=PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"))

    # Capçaleres Taula
    apply_style("A15:F15", "IDENTIFICACIÓ ARTICLE LICITAT", fill=fill_verd, font=bold_9)
    apply_style("G15:R15", "IDENTIFICACIÓ ARTICLE OFERTAT LICITADOR", fill=fill_gris, font=bold_9)
    for c in range(1, 19):
        cell = ws.cell(row=16, column=c)
        cell.font = bold_9; cell.alignment = align_center_wrap; cell.border = border
        if c <= 4 or c == 6: cell.fill = fill_verd
        elif c == 5: cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        else: cell.fill = fill_gris

    # --- PROTECCIÓ I ALÇADES ---
    for r in range(4, 14): ws.row_dimensions[r].height = 25
    ws.row_dimensions[16].height = 45

    last_row = ws.max_row
    for r in range(17, last_row + 1):
        ws.row_dimensions[r].height = None 
        val_cant = ws.cell(row=r, column=5).value
        for c in range(1, 19):
            cell = ws.cell(row=r, column=c)
            cell.alignment = align_center_wrap
            cell.border = border
            if c == 5: cell.number_format = '#,##0'
            cell.protection = Protection(locked=True if val_cant == 0 else (c < 7))

    # Desbloqueig camps d'entrada superiors
    for zona in ["C9:I13", "M9:N12", "O9:R13"]:
        start, end = zona.split(":")
        for row in ws[start:end]:
            for cell in row: cell.protection = Protection(locked=False)

    ws.column_dimensions['D'].width = 55
    ws.protection.password = password_excel
    ws.protection.sheet = True
    
    wb_out.save(nom_sortida)
    print(f"✓ Generat amb èxit: {nom_sortida}")

except Exception as e:
    print(f"Error: {e}")