import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Protection, Border, Side

# 1. Configuración
fitxer_origen = 'hi.xlsm'
nom_sortida = 'ACO2_PPT_AM.xlsx'
password_excel = '1234'

# --- DINAMIC COLUMN MAPPING ---
def get_column_mapping(wb_in, annex_type):
    try:
        cab_file = 'CABECERAS.xlsx'
        df_map = pd.read_excel(cab_file, sheet_name=annex_type, header=None)
        annex_headers = [str(h).strip() for h in df_map.iloc[0].fillna("").tolist()]
        hi_headers_to_find = [str(h).strip().upper() for h in df_map.iloc[1].fillna("").tolist()]
        
        ws_in = wb_in['Full Inici']
        mapping = {}
        
        # Scan headers in rows 5 and 6
        for r_idx in [5, 6]:
            row_vals = [str(ws_in.cell(row=r_idx, column=c).value).strip().upper() for c in range(1, 150)]
            for ah, htf in zip(annex_headers, hi_headers_to_find):
                if htf == "NAN" or htf == "": continue
                if htf in row_vals:
                    mapping[ah] = row_vals.index(htf) + 1 # 1-based column index
        return mapping
    except Exception as e:
        print(f"Error loading mapping: {e}")
        return {}

try:
    print(f"Llegint arxiu d'origen: {fitxer_origen}...")
    # data_only=True per llegir valors calculats de les fórmules
    wb_in = load_workbook(fitxer_origen, data_only=True)
    
    # --- DADES DE CABECERA ---
    nom_pestanya_cab = next((s for s in wb_in.sheetnames if s.upper() == "CABECERA"), None)
    if not nom_pestanya_cab:
        raise ValueError("No s'ha trobat la pestanya 'cabecera' al fitxer original.")
    
    ws_cab = wb_in[nom_pestanya_cab]
    titol_expedient = ws_cab['B9'].value
    num_expedient = ws_cab['B5'].value

    # --- DADES DE PRODUCTES (Full Inici) ---
    ws_in = wb_in['Full Inici']
    
    mapping = get_column_mapping(wb_in, 'AM')
    print(f"Mapeig detectat: {mapping}")

    col_lot = mapping.get("LOTE", 23)
    col_art = mapping.get("ARTÍCULO", 24)
    col_cod = mapping.get("CÓDIGO HCB", 1)
    col_tec = mapping.get("DENOMINACIÓN Y REQUISITOS TÉCNICOS", 32)
    # Busquem també la quantitat de mostres si està mapejada
    col_req = mapping.get("CANTIDAD DE MUESTRAS\nREQUERIDAS", 79)

    dades_extretes = []
    fila_orig = 6
    while ws_in.cell(row=fila_orig, column=col_art).value is not None or ws_in.cell(row=fila_orig, column=col_cod).value is not None:
        val_w = ws_in.cell(row=fila_orig, column=col_lot).value 
        if val_w is not None and str(val_w).strip() != "" and str(val_w).upper() != "NUMERO":
            val_req_val = ws_in.cell(row=fila_orig, column=col_req).value
            try: val_req_val = float(val_req_val) if val_req_val is not None else 1.0
            except: val_req_val = 1.0
            
            dades_extretes.append({
                "lot": val_w, 
                "article": ws_in.cell(row=fila_orig, column=col_art).value,
                "codi_hcb": ws_in.cell(row=fila_orig, column=col_cod).value, 
                "tecnic": ws_in.cell(row=fila_orig, column=col_tec).value,
                "requerides": val_req_val
            })
        fila_orig += 1
        if fila_orig > 10000: break

    # --- CREACIÓ DEL DATAFRAME ---
    cols = ["LOTE", "ARTÍCULO", "CÓDIGO HCB", "DENOMINACIÓN Y REQUISITOS TÉCNICOS", 
            "DENOMINACIÓN ARTÍCULO LICITADOR", "REFERENCIA ARTÍCULO", 
            "CANTIDAD DE MUESTRAS REQUERIDAS", "CANTIDAD DE MUESTRAS PRESENTADAS"]

    df_final = pd.DataFrame(columns=cols)
    if dades_extretes:
        df_temp = pd.DataFrame(dades_extretes)
        df_final["LOTE"] = df_temp["lot"]
        df_final["ARTÍCULO"] = df_temp["article"]
        df_final["CÓDIGO HCB"] = df_temp["codi_hcb"]
        df_final["DENOMINACIÓN Y REQUISITOS TÉCNICOS"] = df_temp["tecnic"]
        df_final["CANTIDAD DE MUESTRAS REQUERIDAS"] = df_temp["requerides"]

    with pd.ExcelWriter(nom_sortida, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, header=False, startrow=15, sheet_name='Plantilla Annex')

    wb_out = load_workbook(nom_sortida)
    ws = wb_out.active

    # --- ESTILS ---
    align_center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
    bold_9 = Font(bold=True, size=9)
    fill_verd = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    fill_gris = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_groc_inst = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    fill_blau = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def apply_style(rng, text=None, fill=None, font=None, align=align_center_wrap):
        if ":" in rng: ws.merge_cells(rng)
        start_cell = ws[rng.split(":")[0]]
        if text is not None: start_cell.value = text
        start_coords, end_coords = rng.split(":") if ":" in rng else (rng, rng)
        for row in ws[start_coords:end_coords]:
            for cell in row:
                cell.alignment = align
                cell.border = border
                if fill: cell.fill = fill
                if font: cell.font = font

    # --- MAQUETACIÓ CABECERA ---
    apply_style("A1:H1", "ANEXO DE CUMPLIMENTACIÓN OBLIGATORIA 2 PPT ALBARÁN DE MUESTRAS", font=Font(bold=True, size=14), fill=fill_blau)
    
    txt_instr = ("Deberá presentar las muestras relacionadas en este albarán según se indique en los pliegos.\n"
                 "Rellene sólo las celdas en blanco o desprotegidas.")
    apply_style("A2:H2", txt_instr, font=Font(italic=True, size=10), fill=fill_groc_inst)
    ws.row_dimensions[2].height = 40

    apply_style("A4:B4", "TÍTULO DEL EXPEDIENTE:", font=bold_9, fill=fill_gris)
    apply_style("C4:H4", titol_expedient, align=align_left_wrap)
    
    apply_style("A5:B5", "NÚMERO DE EXPEDIENTE:", font=bold_9, fill=fill_gris)
    apply_style("C5:H5", num_expedient, align=align_left_wrap)

    apply_style("A7:H7", "LICITADOR E IDENTIFICACIÓN DEL RECEPTOR:", font=bold_9, fill=fill_gris)

    # Bloc Esquerre Licitador
    apply_style("A9:B9", "EMPRESA", fill=fill_gris, font=bold_9); apply_style("C9:D9")
    apply_style("A10:B10", "PERSONA DE CONTACTO", fill=fill_gris, font=bold_9); apply_style("C10:D10")
    apply_style("A11:B11", "TELÉFONO", fill=fill_gris, font=bold_9); apply_style("C11:D11")
    apply_style("A12:B12", "E-MAIL", fill=fill_gris, font=bold_9); apply_style("C12:D12")

    # Bloc Dreta Receptor
    apply_style("E9:F9", "NOMBRE Y DNI/MATRICULA DEL RECEPTOR", fill=fill_gris, font=bold_9, align=align_left_wrap); apply_style("G9:H9")
    apply_style("E10:F10", "FECHA ENTREGA", fill=fill_gris, font=bold_9, align=align_left_wrap); apply_style("G10:H10")
    apply_style("E11:F11", "LUGAR ENTREGA", fill=fill_gris, font=bold_9, align=align_left_wrap); apply_style("G11:H11")

    # Capçaleres Taula
    apply_style("A14:D14", "IDENTIFICACIÓN ARTICULO LICITADO", fill=fill_verd, font=bold_9)
    apply_style("E14:H14", "IDENTIFICACIÓN ARTICULO DEL LICITADOR", fill=fill_gris, font=bold_9)
    
    for c in range(1, 9):
        cell = ws.cell(row=15, column=c)
        cell.font = bold_9; cell.alignment = align_center_wrap; cell.border = border
        if c <= 4 or c == 7: cell.fill = fill_verd
        elif c == 8: cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        else: cell.fill = fill_gris
    
    ws.cell(row=15, column=1, value="LOTE")
    ws.cell(row=15, column=2, value="ARTÍCULO")
    ws.cell(row=15, column=3, value="CÓDIGO HCB")
    ws.cell(row=15, column=4, value="DENOMINACIÓN Y REQUISITOS TÉCNICOS")
    ws.cell(row=15, column=5, value="DENOMINACIÓN ARTÍCULO LICITADOR")
    ws.cell(row=15, column=6, value="REFERENCIA ARTÍCULO")
    ws.cell(row=15, column=7, value="CANTIDAD DE MUESTRAS\nREQUERIDAS")
    ws.cell(row=15, column=8, value="CANTIDAD DE MUESTRAS\nPRESENTADAS")

    # --- PROTECCIÓ I ALÇADES ---
    for r in range(4, 13): ws.row_dimensions[r].height = 25
    ws.row_dimensions[14].height = 30
    ws.row_dimensions[15].height = 65

    last_row = ws.max_row
    for r in range(16, last_row + 1):
        ws.row_dimensions[r].height = None 
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.alignment = align_center_wrap
            cell.border = border
            # Lote (1), Articulo (2), Codigo (3), Denominacion tecnica (4), y Cantidad Requerida (7) no se tocan.
            # Denominacion Licitador (5), Referencia (6), Cantidad Presentada (8) si.
            if c in [1, 2, 3, 4, 7]:
                cell.protection = Protection(locked=True)
            else:
                cell.protection = Protection(locked=False)

    # Desbloqueig camps d'entrada superiors
    # C9:D12 Licitador, G9:H11 Receptor
    for zona in ["C9:D12", "G9:H11"]:
        start, end = zona.split(":")
        for row in ws[start:end]:
            for cell in row: cell.protection = Protection(locked=False)

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    
    ws.protection.password = password_excel
    ws.protection.sheet = True
    
    wb_out.save(nom_sortida)
    print(f"✓ Generat amb èxit: {nom_sortida}")

except Exception as e:
    print(f"Error: {e}")
